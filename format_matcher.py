"""
format_matcher.py

"Format Matcher" feature: takes a reference/template Excel file — whose
formatting (fonts, fills, borders, column widths, row heights, merged
cells, number formats, freeze panes) defines the exact layout you want —
and pastes another file's raw data into it, cloning that formatting onto
every row of pasted data.

The reference file IS the template: we open it as-is and only write
values into the data rows. Whatever formatting already sits on the
"style source" row (see apply_template) gets stamped onto every row of
pasted data, including rows added beyond what the template originally had.
"""
import os
import csv
import copy
import datetime
import openpyxl


def _coerce_csv_value(text):
    text = text.strip()
    if text == "":
        return None
    for caster in (int, float):
        try:
            return caster(text)
        except ValueError:
            continue
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return text


def read_source_rows(data_path, sheet_name=None, has_header=True):
    """Read raw values from the data file, preserving types (numbers,
    dates, strings) as closely as possible. Returns a list of rows, each
    a list of cell values, with the header row stripped if present.
    Blank rows are skipped."""
    ext = os.path.splitext(data_path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        wb = openpyxl.load_workbook(data_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                rows.append(list(row))
        if has_header and rows:
            rows = rows[1:]
        return rows
    elif ext == ".csv":
        rows = []
        with open(data_path, newline="", encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if any(cell.strip() for cell in row):
                    rows.append([_coerce_csv_value(cell) for cell in row])
        if has_header and rows:
            rows = rows[1:]
        return rows
    else:
        raise ValueError(f"Unsupported data file type: {ext}")


def list_sheet_names(path):
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    return wb.sheetnames


def _copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy.copy(src_cell.protection)


def _merged_spans_starting_at_row(ws, row_idx):
    """Column spans (start_col, end_col) of any merged range that begins
    on this row, so the same merge pattern can be replicated on rows
    added beyond the template's original extent."""
    spans = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == row_idx:
            spans.append((merged_range.min_col, merged_range.max_col))
    return spans


def apply_template(reference_path, data_path, output_path,
                    style_source_row=2, max_column=None,
                    reference_sheet_name=None, data_sheet_name=None,
                    data_has_header=True, progress_cb=None, log_cb=None):
    """
    reference_path: the template .xlsx — its formatting is preserved exactly.
    data_path: the .xlsx/.xlsm/.csv with the raw data to paste in.
    output_path: where to save the result (a filled-in copy of the template).
    style_source_row: the template row whose styling (font/fill/border/
        alignment/number format/row height/merge pattern) gets stamped onto
        every data row, including new rows added past the template's
        original size. Default 2 assumes row 1 is the header.
    max_column: how many columns to write per row. Defaults to however
        many columns style_source_row already spans in the template.
    progress_cb(fraction): optional, called after each row is written.
    log_cb(message): optional, called with human-readable status lines.
    """
    def log(msg):
        if log_cb:
            log_cb(msg)

    log(f"Loading template: {os.path.basename(reference_path)}")
    wb = openpyxl.load_workbook(reference_path)
    ws = wb[reference_sheet_name] if reference_sheet_name else wb.active

    if style_source_row < 1:
        raise ValueError("style_source_row must be 1 or greater")
    if max_column is None:
        max_column = max(ws.max_column, 1)

    style_cells = [ws.cell(row=style_source_row, column=c) for c in range(1, max_column + 1)]
    style_row_height = ws.row_dimensions[style_source_row].height
    merge_spans = _merged_spans_starting_at_row(ws, style_source_row)

    log(f"Reading data: {os.path.basename(data_path)}")
    rows = read_source_rows(data_path, sheet_name=data_sheet_name, has_header=data_has_header)
    total = len(rows)
    log(f"{total} data row(s) to paste into the template, starting at row {style_source_row}.")

    for i, row_values in enumerate(rows):
        target_row = style_source_row + i
        for col in range(1, max_column + 1):
            value = row_values[col - 1] if col - 1 < len(row_values) else None
            cell = ws.cell(row=target_row, column=col, value=value)
            _copy_cell_style(style_cells[col - 1], cell)
        if style_row_height is not None:
            ws.row_dimensions[target_row].height = style_row_height
        if target_row != style_source_row:
            for start_col, end_col in merge_spans:
                try:
                    ws.merge_cells(start_row=target_row, start_column=start_col,
                                   end_row=target_row, end_column=end_col)
                except Exception:
                    pass  # range already merged on a reused template row — fine
        if progress_cb and total:
            progress_cb((i + 1) / total)

    log(f"Saving: {os.path.basename(output_path)}")
    wb.save(output_path)
    log("Done.")
    return output_path
