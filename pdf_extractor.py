"""
pdf_extractor.py
Backend logic for the "PDF Extractor" feature: reads text out of PDFs,
sends it to OpenAI (gpt-4o) with a user-supplied prompt, and appends
the structured result as a new row in an Excel sheet.

Kept separate from the UI so it can be tested/reused on its own,
matching the pattern used by downloader.py.
"""
import os
import time
import threading

import pdfplumber
from openpyxl import Workbook, load_workbook

try:
    from openai import OpenAI, AuthenticationError, APIError, APITimeoutError, APIConnectionError
except ImportError:  # openai package not installed yet — handled gracefully at call time
    OpenAI = None
    AuthenticationError = APIError = APITimeoutError = APIConnectionError = Exception

ENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 2  # doubles each retry: 2s, 4s, 8s...


# ---------------- API key storage (.env) ----------------

def save_api_key(api_key: str):
    """Writes/overwrites OPENAI_API_KEY in a local .env file (never committed — see .gitignore)."""
    lines = []
    found = False
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH, "r", encoding="utf-8") as f:
            for line in f:
                if line.startswith("OPENAI_API_KEY="):
                    lines.append(f"OPENAI_API_KEY={api_key}\n")
                    found = True
                else:
                    lines.append(line)
    if not found:
        lines.append(f"OPENAI_API_KEY={api_key}\n")
    with open(ENV_PATH, "w", encoding="utf-8") as f:
        f.writelines(lines)


def load_api_key() -> str:
    """Returns the saved API key, or '' if none is stored yet."""
    if not os.path.exists(ENV_PATH):
        return ""
    with open(ENV_PATH, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("OPENAI_API_KEY="):
                return line.split("=", 1)[1].strip()
    return ""


# ---------------- PDF text extraction ----------------

def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Returns all text from a PDF, page by page. Raises ValueError with a
    clear, human-readable message on failure (encrypted, corrupted, or
    unreadable file) instead of letting a blank/cryptic exception through.
    """
    try:
        text_parts = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text_parts.append(page_text)
        text = "\n".join(text_parts).strip()
        if not text:
            raise ValueError("No extractable text found (the PDF may be scanned/image-only).")
        return text
    except ValueError:
        raise
    except Exception as e:
        msg = str(e).strip()
        if not msg:
            msg = "the file may be password-protected or corrupted"
        raise ValueError(f"Could not read PDF: {msg}")


# ---------------- OpenAI call (with retry/backoff) ----------------

def call_openai(client, prompt: str, pdf_text: str, model: str = "gpt-4o"):
    """
    Sends prompt + extracted PDF text to OpenAI. Retries on timeout/
    connection errors with exponential backoff. Raises AuthenticationError
    immediately (no point retrying a bad key) and ValueError for other
    unrecoverable API errors after retries are exhausted.
    """
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You extract structured information from documents."},
                    {"role": "user", "content": f"{prompt}\n\n---\nDocument text:\n{pdf_text}"},
                ],
            )
            return response.choices[0].message.content.strip()
        except AuthenticationError:
            raise  # bad key — don't waste retries, let caller short-circuit the batch
        except (APITimeoutError, APIConnectionError) as e:
            last_error = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_SECONDS * (2 ** (attempt - 1)))
                continue
            raise ValueError(f"OpenAI request timed out after {MAX_RETRIES} attempts: {e}")
        except APIError as e:
            raise ValueError(f"OpenAI API error: {e}")
    raise ValueError(f"OpenAI request failed: {last_error}")


# ---------------- Excel append ----------------

def append_results_to_excel(save_path: str, rows: list, headers: list = None):
    """
    Appends rows to save_path (creates the file + header row if it
    doesn't exist yet). Never overwrites existing rows — each run's
    results stack underneath the previous ones.
    rows: list of tuples/lists, one per PDF processed.
    """
    headers = headers or ["File Name", "Extracted Result"]
    if os.path.exists(save_path):
        wb = load_workbook(save_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    for row in rows:
        ws.append(row)

    wb.save(save_path)


# ---------------- Threaded pipeline ----------------

def process_pdfs(pdf_paths: list, api_key: str, prompt: str, excel_save_path: str,
                  progress_cb=None, log_cb=None, model: str = "gpt-4o"):
    """
    Processes each PDF independently (extract text -> call OpenAI ->
    collect result). Writes all successful results to excel_save_path
    in one batch at the end.

    progress_cb(done, total) — called after each PDF finishes (success or fail).
    log_cb(message) — called for every warning/error.

    Short-circuits the WHOLE batch immediately if the API key is invalid
    (no point burning through every file with a key that will never work).

    Returns (ok_count, failed_count).
    """
    progress_cb = progress_cb or (lambda done, total: None)
    log_cb = log_cb or (lambda msg: None)

    if not api_key or not api_key.strip():
        log_cb("No OpenAI API key set. Please enter your key first.")
        return 0, len(pdf_paths)

    if OpenAI is None:
        log_cb("The 'openai' package isn't installed. Run: pip install openai")
        return 0, len(pdf_paths)

    client = OpenAI(api_key=api_key)

    total = len(pdf_paths)
    done = 0
    ok = 0
    failed = 0
    results = []
    lock = threading.Lock()
    abort_event = threading.Event()

    def worker(pdf_path):
        nonlocal done, ok, failed
        if abort_event.is_set():
            return
        file_name = os.path.basename(pdf_path)
        try:
            text = extract_text_from_pdf(pdf_path)
            result_text = call_openai(client, prompt, text, model=model)
            with lock:
                results.append((file_name, result_text))
                ok += 1
        except AuthenticationError:
            abort_event.set()
            log_cb("Invalid OpenAI API key — stopping the batch.")
            with lock:
                failed += 1
        except ValueError as e:
            log_cb(f"{file_name}: {e}")
            with lock:
                failed += 1
        finally:
            with lock:
                done += 1
                progress_cb(done, total)

    threads = []
    for pdf_path in pdf_paths:
        if abort_event.is_set():
            break
        t = threading.Thread(target=worker, args=(pdf_path,), daemon=True)
        threads.append(t)
        t.start()

    for t in threads:
        t.join()

    if results:
        append_results_to_excel(excel_save_path, results)

    # Anything left un-attempted because of an auth short-circuit counts as failed.
    failed += (total - done)

    return ok, failed
