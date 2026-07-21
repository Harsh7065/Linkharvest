"""
dashboard_builder.py
Backend logic for the "Dashboard Builder" sidebar page in LinkHarvest.

Given any CSV/Excel file, this module:
1. Classifies every column (numeric / categorical / datetime / identifier /
   text) using dtype + simple cardinality heuristics.
2. Proposes a chart plan — which chart type fits which column(s), and why —
   using well-established data-viz rules of thumb (e.g. a trend over time
   is a line chart, a category breakdown with few groups is a bar/pie
   chart, a numeric column's spread is a histogram).
3. Renders a quick on-screen preview (matplotlib figure) so the plan can be
   sanity-checked before committing to anything.
4. Builds a real, editable Excel workbook: a "Data" sheet plus a
   "Dashboard" sheet with native openpyxl chart objects laid out in a grid
   — this is a *reference* dashboard the user can restyle/extend in Excel.
5. Optionally asks an AI provider (same OpenAI/Gemini engine + saved key
   already used by the PDF Extractor / Excel Editor pages) to write a
   short plain-English summary explaining the logic behind the structure.

No UI code lives here. Keep this importable and testable on its own.
"""
import os
import json
import time

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows

import pdf_extractor as pe  # reuse provider constants, saved .env keys, retry timeouts

try:
    from openai import OpenAI
    from openai import (
        AuthenticationError as OpenAIAuthenticationError,
        APITimeoutError as OpenAIAPITimeoutError,
        RateLimitError as OpenAIRateLimitError,
        APIError as OpenAIAPIError,
    )
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import google.generativeai as genai
    from google.api_core.exceptions import (
        GoogleAPIError,
        InvalidArgument as GoogleInvalidArgument,
        PermissionDenied as GooglePermissionDenied,
        ResourceExhausted as GoogleResourceExhausted,
        DeadlineExceeded as GoogleDeadlineExceeded,
    )
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False


class AuthError(Exception):
    """Raised when the selected provider's API key is missing or invalid."""


MAX_CATEGORY_UNIQUES = 15      # above this, a column is treated as "high-cardinality" (id-like/text)
MAX_PIE_SLICES = 8             # beyond this a bar chart reads better than a pie
MAX_CHARTS_IN_PLAN = 8         # keep the reference dashboard readable
MAX_RETRIES = 3
RETRY_BACKOFF_SECONDS = 2
API_TIMEOUT = 60


# ==========================================================================
# 1. Column classification
# ==========================================================================
def classify_columns(df: pd.DataFrame) -> list:
    """
    Returns a list of {"column", "kind", "unique_count", "missing"} dicts.
    kind is one of: "datetime", "numeric", "categorical", "identifier", "text".
    """
    info = []
    n = max(len(df), 1)
    for col in df.columns:
        series = df[col]
        unique_count = int(series.nunique(dropna=True))
        missing = int(series.isna().sum())

        if pd.api.types.is_datetime64_any_dtype(series):
            kind = "datetime"
        else:
            # Try a light datetime parse on text-like columns (e.g. "2024-01-05" strings).
            # Covers both legacy object dtype and pandas' newer dedicated string dtype.
            parsed_dt = None
            if pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series):
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", UserWarning)
                    parsed_dt = pd.to_datetime(series, errors="coerce", format=None)
                if parsed_dt.notna().mean() > 0.8:
                    kind = "datetime"
                else:
                    parsed_dt = None
            if parsed_dt is None:
                if pd.api.types.is_numeric_dtype(series):
                    # A numeric column with very few distinct values behaves
                    # more like a category (e.g. a 1-5 rating or a status code).
                    kind = "categorical" if unique_count <= min(10, n) and unique_count < n * 0.05 + 5 else "numeric"
                elif n >= 10 and unique_count >= n * 0.9:
                    # Almost every value is unique — an ID/key column, not a
                    # meaningful category to chart. Checked before the flat
                    # cardinality cutoff so small ID columns aren't missed.
                    kind = "identifier"
                elif unique_count <= MAX_CATEGORY_UNIQUES:
                    kind = "categorical"
                else:
                    kind = "text"

        info.append({
            "column": str(col),
            "kind": kind,
            "unique_count": unique_count,
            "missing": missing,
        })
    return info


# ==========================================================================
# 2. Chart plan — which chart fits which column(s), and why
# ==========================================================================
def suggest_chart_plan(df: pd.DataFrame, column_info: list = None) -> list:
    """
    Returns an ordered list of chart suggestions:
        {"title", "chart_type", "x", "y", "reason"}
    chart_type is one of: "line", "bar", "pie", "histogram", "scatter".
    Capped at MAX_CHARTS_IN_PLAN so the reference dashboard stays readable.
    """
    column_info = column_info if column_info is not None else classify_columns(df)
    by_kind = {}
    for c in column_info:
        by_kind.setdefault(c["kind"], []).append(c)

    datetimes = by_kind.get("datetime", [])
    numerics = by_kind.get("numeric", [])
    categoricals = by_kind.get("categorical", [])

    plan = []

    # Trend over time: first datetime column x each numeric column (line chart)
    if datetimes and numerics:
        dt_col = datetimes[0]["column"]
        for num in numerics[:3]:
            plan.append({
                "title": f"{num['column']} over {dt_col}",
                "chart_type": "line",
                "x": dt_col, "y": num["column"],
                "reason": (f"'{dt_col}' is a date/time column and '{num['column']}' is numeric — "
                           "a line chart is the standard way to show a trend over time."),
            })

    # Category breakdown: each low-cardinality categorical vs first numeric
    # (bar for the comparison), and value-count pie/bar for its own distribution.
    for cat in categoricals[:3]:
        if numerics:
            num = numerics[0]["column"]
            plan.append({
                "title": f"{num} by {cat['column']}",
                "chart_type": "bar",
                "x": cat["column"], "y": num,
                "reason": (f"'{cat['column']}' has only {cat['unique_count']} distinct groups — "
                           f"a bar chart compares '{num}' cleanly across each group."),
            })
        chart_type = "pie" if cat["unique_count"] <= MAX_PIE_SLICES else "bar"
        plan.append({
            "title": f"Share of records by {cat['column']}",
            "chart_type": chart_type,
            "x": cat["column"], "y": None,
            "reason": (f"Shows how records are distributed across '{cat['column']}''s "
                       f"{cat['unique_count']} categories — "
                       + ("a pie works since there are few slices." if chart_type == "pie"
                          else "a bar chart instead of a pie since there are too many slices to read as wedges.")),
        })

    # Distribution: histogram for numeric columns without a natural category/time pairing
    for num in numerics[:3]:
        plan.append({
            "title": f"Distribution of {num['column']}",
            "chart_type": "histogram",
            "x": num["column"], "y": None,
            "reason": (f"'{num['column']}' is numeric with {num['unique_count']} distinct values — "
                       "a histogram shows its spread and where most values cluster."),
        })

    # Correlation: first two numeric columns, if there are at least two
    if len(numerics) >= 2:
        x, y = numerics[0]["column"], numerics[1]["column"]
        plan.append({
            "title": f"{y} vs {x}",
            "chart_type": "scatter",
            "x": x, "y": y,
            "reason": (f"Both '{x}' and '{y}' are numeric — a scatter plot is the standard way "
                       "to check whether the two move together."),
        })

    return plan[:MAX_CHARTS_IN_PLAN]


# ==========================================================================
# 2b. SQL/DAX-style aggregates — real computed numbers, not just chart
# metadata, so the AI summary/plan reasoning is grounded in the actual data
# (equivalent to a GROUP BY ... SUM/AVG/COUNT, or a DAX measure).
# ==========================================================================
def compute_group_aggregates(df: pd.DataFrame, column_info: list = None, top_n: int = 5) -> list:
    """
    For each categorical column paired with each numeric column, computes
    the pandas equivalent of:
        SELECT category, SUM(value), AVG(value), COUNT(*)
        FROM df GROUP BY category ORDER BY SUM(value) DESC LIMIT top_n
    Returns a list of {"category_column", "value_column", "top_groups": [...]}
    dicts, capped to keep the payload sent to the AI small.
    """
    column_info = column_info if column_info is not None else classify_columns(df)
    categoricals = [c["column"] for c in column_info if c["kind"] == "categorical"]
    numerics = [c["column"] for c in column_info if c["kind"] == "numeric"]

    aggregates = []
    for cat in categoricals[:3]:
        for num in numerics[:2]:
            try:
                grouped = (
                    df.groupby(cat)[num]
                    .agg(sum="sum", avg="mean", count="count")
                    .sort_values("sum", ascending=False)
                    .head(top_n)
                )
            except Exception:
                continue
            aggregates.append({
                "category_column": cat,
                "value_column": num,
                "top_groups": [
                    {"group": str(idx), "sum": round(float(row["sum"]), 2),
                     "avg": round(float(row["avg"]), 2), "count": int(row["count"])}
                    for idx, row in grouped.iterrows()
                ],
            })
    return aggregates


# ==========================================================================
# 3. On-screen preview (matplotlib grid, embedded via FigureCanvasTkAgg)
# ==========================================================================
_BG = "#1F1F1F"
_TEXT = "#E6E6E6"
_PALETTE = ["#4DA3FF", "#FF7A45", "#8C8C8C", "#E8A33D", "#C77DFF", "#FF5D8F", "#3ca34d", "#e05050"]


def render_dashboard_preview(df: pd.DataFrame, plan: list, figsize=(11, 7)):
    """Builds a matplotlib Figure with one subplot per chart in the plan."""
    if not plan:
        fig, ax = plt.subplots(figsize=figsize, facecolor=_BG)
        ax.set_facecolor(_BG)
        ax.text(0.5, 0.5, "Not enough structured columns to suggest charts",
                ha="center", va="center", color=_TEXT, fontsize=12, transform=ax.transAxes)
        ax.axis("off")
        return fig

    n = len(plan)
    cols = 2 if n > 1 else 1
    rows = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows, cols, figsize=figsize, facecolor=_BG)
    axes = axes.flatten() if n > 1 else [axes]

    for i, spec in enumerate(plan):
        ax = axes[i]
        ax.set_facecolor(_BG)
        try:
            _render_one_chart(ax, df, spec, i)
        except Exception as e:
            ax.text(0.5, 0.5, f"Couldn't render:\n{e}", ha="center", va="center",
                     color="#e05050", fontsize=9, transform=ax.transAxes, wrap=True)
        ax.set_title(spec["title"], color=_TEXT, fontsize=10)
        ax.tick_params(colors=_TEXT, labelsize=7)
        for spine in ax.spines.values():
            spine.set_color("#444444")

    for j in range(n, len(axes)):
        axes[j].axis("off")

    fig.tight_layout()
    return fig


def _render_one_chart(ax, df, spec, color_idx):
    color = _PALETTE[color_idx % len(_PALETTE)]
    ctype = spec["chart_type"]

    if ctype == "line":
        x_vals = pd.to_datetime(df[spec["x"]], errors="coerce")
        y_vals = pd.to_numeric(df[spec["y"]], errors="coerce")
        ordered = pd.DataFrame({"x": x_vals, "y": y_vals}).dropna().sort_values("x")
        ax.plot(ordered["x"], ordered["y"], color=color, linewidth=1.6)

    elif ctype == "bar" and spec["y"]:
        grouped = df.groupby(spec["x"])[spec["y"]].mean(numeric_only=True).sort_values(ascending=False).head(12)
        ax.bar(grouped.index.astype(str), grouped.values, color=color)
        ax.tick_params(axis="x", rotation=45)

    elif ctype == "bar":
        counts = df[spec["x"]].value_counts().head(12)
        ax.bar(counts.index.astype(str), counts.values, color=color)
        ax.tick_params(axis="x", rotation=45)

    elif ctype == "pie":
        counts = df[spec["x"]].value_counts().head(MAX_PIE_SLICES)
        ax.pie(counts.values, labels=counts.index.astype(str), colors=_PALETTE,
               textprops={"color": _TEXT, "fontsize": 7}, autopct="%1.0f%%")

    elif ctype == "histogram":
        vals = pd.to_numeric(df[spec["x"]], errors="coerce").dropna()
        ax.hist(vals, bins=min(20, max(5, vals.nunique())), color=color)

    elif ctype == "scatter":
        x_vals = pd.to_numeric(df[spec["x"]], errors="coerce")
        y_vals = pd.to_numeric(df[spec["y"]], errors="coerce")
        ordered = pd.DataFrame({"x": x_vals, "y": y_vals}).dropna()
        ax.scatter(ordered["x"], ordered["y"], color=color, s=10, alpha=0.7)


# ==========================================================================
# 4. Real Excel dashboard (Data sheet + Dashboard sheet with native charts)
# ==========================================================================
def build_excel_dashboard(df: pd.DataFrame, plan: list, output_path: str, summary_text: str = ""):
    """
    Writes a workbook with:
      - "Data" sheet: the raw data, used as the source range for every chart
      - "Dashboard" sheet: native openpyxl charts laid out in a 2-column grid
      - "Summary" sheet: the plain-English explanation of the structure (if provided)
    The result is a real, editable .xlsx — every chart stays live and
    re-styleable in Excel, not a flattened image.
    """
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Data"
    for row in dataframe_to_rows(df, index=False, header=True):
        data_ws.append(row)
    max_row = data_ws.max_row
    headers = [str(c) for c in df.columns]

    dash_ws = wb.create_sheet("Dashboard")
    dash_ws.sheet_view.showGridLines = False
    dash_ws["A1"] = "Dashboard Reference"
    dash_ws["A1"].font = dash_ws["A1"].font.copy(bold=True, size=16)

    col_positions = ["B", "N"]   # two-column chart grid
    row_cursor = 3
    row_height_per_chart = 16

    for i, spec in enumerate(plan):
        chart = _build_excel_chart(spec, data_ws, headers, max_row)
        if chart is None:
            continue
        anchor_col = col_positions[i % 2]
        anchor = f"{anchor_col}{row_cursor}"
        dash_ws.add_chart(chart, anchor)
        if i % 2 == 1:
            row_cursor += row_height_per_chart

    if summary_text:
        sum_ws = wb.create_sheet("Summary")
        sum_ws["A1"] = "Dashboard Logic — Summary"
        sum_ws["A1"].font = sum_ws["A1"].font.copy(bold=True, size=14)
        sum_ws.column_dimensions["A"].width = 110
        for i, line in enumerate(summary_text.split("\n"), start=3):
            cell = sum_ws.cell(row=i, column=1, value=line)
            cell.alignment = cell.alignment.copy(wrap_text=True)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def read_reference_text(reference_path: str, max_chars: int = 4000) -> str:
    """
    Reads an optional user-supplied 'reference' file describing the kind of
    dashboard they want. Plain text/markdown files are read as-is; a
    CSV/Excel reference is summarized (columns + a few rows) rather than
    dumped whole. Returns "" if the file can't be read — this is always
    optional, so a read failure should never block dashboard generation.
    """
    if not reference_path or not os.path.isfile(reference_path):
        return ""
    ext = os.path.splitext(reference_path)[1].lower()
    try:
        if ext in (".csv", ".xlsx", ".xls"):
            df = pd.read_csv(reference_path) if ext == ".csv" else pd.read_excel(reference_path)
            text = (f"Reference file columns: {', '.join(map(str, df.columns))}\n"
                    f"Sample rows:\n{df.head(5).to_csv(index=False)}")
        else:
            with open(reference_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        return text[:max_chars]
    except Exception:
        return ""


# ==========================================================================
# 4b. Power BI-ready export — a proper Excel Table (ListObject), which
# Power BI's "Get Data > Excel Workbook" reads directly as a query-ready
# table (unlike a plain, un-tabled data range). We can't generate an actual
# .pbix file outside Power BI itself, so this is the standard hand-off
# format plus a short sheet explaining the one-click import.
# ==========================================================================
def export_power_bi_ready(df: pd.DataFrame, output_path: str) -> str:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    n_rows, n_cols = ws.max_row, ws.max_column
    last_col_letter = ws.cell(row=1, column=n_cols).column_letter
    table_ref = f"A1:{last_col_letter}{n_rows}"
    table = Table(displayName="DashboardData", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False)
    ws.add_table(table)

    howto = wb.create_sheet("Import to Power BI")
    howto["A1"] = "Importing this workbook into Power BI"
    howto["A1"].font = howto["A1"].font.copy(bold=True, size=14)
    steps = [
        "1. In Power BI Desktop: Home > Get Data > Excel Workbook.",
        f"2. Select this file, then choose the 'DashboardData' table on the 'Data' sheet "
        "(it's a real Excel Table, so Power BI reads it as a clean query-ready source).",
        "3. Click Load (or Transform Data first if you want to reshape columns in Power Query).",
        "4. Build visuals from the loaded table as usual.",
        "",
        "Note: LinkHarvest can't generate a .pbix file directly — Power BI Desktop only "
        "writes that format itself — but this Table-based .xlsx is the standard, "
        "zero-friction hand-off format Power BI expects.",
    ]
    for i, line in enumerate(steps, start=3):
        cell = howto.cell(row=i, column=1, value=line)
        cell.alignment = cell.alignment.copy(wrap_text=True)
    howto.column_dimensions["A"].width = 100

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def _col_letter_index(headers, name):
    return headers.index(name) + 1  # openpyxl columns are 1-indexed


def _build_excel_chart(spec, data_ws, headers, max_row):
    ctype = spec["chart_type"]
    try:
        if ctype == "line":
            chart = LineChart()
            chart.title = spec["title"]
            y_idx = _col_letter_index(headers, spec["y"])
            x_idx = _col_letter_index(headers, spec["x"])
            data_ref = Reference(data_ws, min_col=y_idx, min_row=1, max_row=max_row)
            cats_ref = Reference(data_ws, min_col=x_idx, min_row=2, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

        elif ctype == "bar":
            chart = BarChart()
            chart.title = spec["title"]
            chart.type = "col"
            y_name = spec["y"] or headers[0]
            y_idx = _col_letter_index(headers, y_name) if spec["y"] else None
            x_idx = _col_letter_index(headers, spec["x"])
            if y_idx:
                data_ref = Reference(data_ws, min_col=y_idx, min_row=1, max_row=max_row)
                chart.add_data(data_ref, titles_from_data=True)
            cats_ref = Reference(data_ws, min_col=x_idx, min_row=2, max_row=max_row)
            chart.set_categories(cats_ref)

        elif ctype == "pie":
            chart = PieChart()
            chart.title = spec["title"]
            x_idx = _col_letter_index(headers, spec["x"])
            data_ref = Reference(data_ws, min_col=x_idx, min_row=1, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)

        elif ctype == "histogram":
            chart = BarChart()
            chart.title = spec["title"] + " (raw values — group in Excel for true bins)"
            chart.type = "col"
            x_idx = _col_letter_index(headers, spec["x"])
            data_ref = Reference(data_ws, min_col=x_idx, min_row=1, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)

        elif ctype == "scatter":
            chart = ScatterChart()
            chart.title = spec["title"]
            x_idx = _col_letter_index(headers, spec["x"])
            y_idx = _col_letter_index(headers, spec["y"])
            xref = Reference(data_ws, min_col=x_idx, min_row=2, max_row=max_row)
            yref = Reference(data_ws, min_col=y_idx, min_row=1, max_row=max_row)
            series = Series(yref, xref, title_from_data=True)
            chart.series.append(series)
        else:
            return None

        chart.width = 16
        chart.height = 9
        return chart
    except (ValueError, KeyError):
        return None


# ==========================================================================
# 5. Optional AI summary of the dashboard logic (reuses pe's provider infra)
# ==========================================================================
def _build_summary_prompt(column_info, plan, aggregates=None, reference_text=None) -> str:
    cols_desc = "\n".join(
        f"- {c['column']}: {c['kind']} ({c['unique_count']} unique values, {c['missing']} missing)"
        for c in column_info
    )
    charts_desc = "\n".join(
        f"- {p['title']} ({p['chart_type']}): {p['reason']}" for p in plan
    )
    agg_desc = ""
    if aggregates:
        lines = []
        for a in aggregates:
            top = ", ".join(f"{g['group']}={g['sum']}" for g in a["top_groups"][:3])
            lines.append(f"- SUM({a['value_column']}) GROUP BY {a['category_column']}: {top}")
        agg_desc = "\n\nActual computed aggregates (GROUP BY-style, use these real numbers):\n" + "\n".join(lines)

    ref_desc = ""
    if reference_text:
        ref_desc = (
            "\n\nThe user described the kind of dashboard/insight they want (use this to shape "
            f"which points you emphasize, but stay grounded in the actual data above):\n{reference_text}"
        )

    return (
        "You are a data analyst explaining a dashboard's structure to someone who has "
        "never built one before. Given this column breakdown:\n"
        f"{cols_desc}\n\n"
        "And this proposed set of charts:\n"
        f"{charts_desc}"
        f"{agg_desc}"
        f"{ref_desc}\n\n"
        "Write a short, plain-English summary (5-9 sentences, no headings, no markdown) that: "
        "explains WHY this data was grouped into these particular chart types, what each chart "
        "tells the reader at a glance (citing real numbers from the aggregates when available), "
        "and one or two things to watch out for (e.g. skewed categories, missing data) when "
        "interpreting it. Do not repeat the raw lists back verbatim."
    )


def _call_openai_summary(client, model_name, prompt):
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
                timeout=API_TIMEOUT,
            )
            return response.choices[0].message.content.strip()
        except OpenAIAuthenticationError as e:
            raise AuthError(f"Invalid or missing OpenAI API key: {e}")
        except (OpenAIAPITimeoutError, OpenAIRateLimitError) as e:
            last_error = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_SECONDS * (2 ** (attempt - 1)))
                continue
            raise RuntimeError(f"OpenAI timed out/rate-limited after {MAX_RETRIES} attempts: {e}")
        except OpenAIAPIError as e:
            raise RuntimeError(f"OpenAI API error (check the model name '{model_name}' is valid): {e}")
    raise RuntimeError(f"Summary generation failed: {last_error}")


def _call_gemini_summary(model, prompt):
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = model.generate_content(
                prompt,
                generation_config=genai.types.GenerationConfig(temperature=0.3),
                request_options={"timeout": API_TIMEOUT},
            )
            return response.text.strip()
        except GooglePermissionDenied as e:
            raise AuthError(f"Invalid or missing Gemini API key: {e}")
        except GoogleInvalidArgument as e:
            msg = str(e).lower()
            if "api key" in msg or "api_key" in msg:
                raise AuthError(f"Invalid Gemini API key: {e}")
            raise RuntimeError(f"Gemini rejected the request: {e}")
        except (GoogleResourceExhausted, GoogleDeadlineExceeded) as e:
            last_error = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_BACKOFF_SECONDS * (2 ** (attempt - 1)))
                continue
            raise RuntimeError(f"Gemini timed out/rate-limited after {MAX_RETRIES} attempts: {e}")
        except GoogleAPIError as e:
            raise RuntimeError(f"Gemini API error: {e}")
    raise RuntimeError(f"Summary generation failed: {last_error}")


def generate_ai_summary(column_info, plan, provider: str, api_key: str, model_name: str = None,
                         aggregates=None, reference_text: str = None) -> str:
    """
    Returns a short plain-English explanation of the dashboard's logic.
    aggregates: output of compute_group_aggregates(), grounds the summary in
        real GROUP BY-style numbers instead of just chart metadata.
    reference_text: optional free-text description of the kind of dashboard
        the user wants (e.g. pasted from a reference file) — used to steer
        emphasis, never required.
    Raises AuthError / RuntimeError on failure — caller decides whether a
    missing/failed summary should block the rest of the export (it shouldn't;
    the charts and Excel file are useful without it).
    """
    if provider not in (pe.PROVIDER_OPENAI, pe.PROVIDER_GEMINI):
        raise ValueError(f"Unknown provider: {provider}")
    if not is_provider_available(provider):
        pkg = "openai" if provider == pe.PROVIDER_OPENAI else "google-generativeai"
        raise RuntimeError(f"The '{pkg}' package isn't installed. Run: pip install {pkg}")
    if not api_key:
        raise AuthError(f"No {provider.title()} API key configured.")

    model_name = (model_name or "").strip()
    if not model_name:
        model_name = pe.DEFAULT_OPENAI_MODEL if provider == pe.PROVIDER_OPENAI else pe.DEFAULT_GEMINI_MODEL

    prompt = _build_summary_prompt(column_info, plan, aggregates=aggregates, reference_text=reference_text)

    if provider == pe.PROVIDER_OPENAI:
        client = OpenAI(api_key=api_key)
        return _call_openai_summary(client, model_name, prompt)
    else:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(model_name)
        return _call_gemini_summary(model, prompt)


def is_provider_available(provider: str) -> bool:
    return OPENAI_AVAILABLE if provider == pe.PROVIDER_OPENAI else GEMINI_AVAILABLE
