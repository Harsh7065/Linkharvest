"""
downloader.py
Core scanning / downloading logic, adapted from the original
Download_All_Row_Images.py script. Kept separate from the UI so it
can be tested or reused on its own.
"""
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook


def load_sheet(workbook_path: str, sheet_name: str):
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook!")
    return wb[sheet_name]


def guess_extension(url: str) -> str:
    """Best-effort file extension based on the URL, so videos/audio
    don't get saved with a wrong .jpg extension."""
    lower = url.lower().split("?")[0]
    for ext in (".jpg", ".jpeg", ".png", ".gif", ".webp",
                ".mp4", ".mov", ".avi", ".mp3", ".wav", ".m4a"):
        if lower.endswith(ext):
            return ext
    return ".jpg"


def build_download_tasks(ws, row_from: int, row_to: int, col_from: int,
                          col_to: int, folder_path: str):
    """
    Scan rows row_from..row_to and columns col_from..col_to (inclusive)
    for any cell containing a URL ("http" substring), exactly like the
    original VBA logic. Column A of each row is always used as the ID
    for naming files, regardless of the chosen column range.

    Pass row_from == row_to for a single specific row, and/or
    col_from == col_to for a single specific column.
    """
    tasks = []
    max_row = ws.max_row
    max_col = ws.max_column

    row_from = max(2, row_from)
    row_to = min(max_row, row_to) if row_to else max_row
    col_from = max(1, col_from)
    col_to = min(max_col, col_to) if col_to else max_col

    for i in range(row_from, row_to + 1):
        id_value = ws.cell(row=i, column=1).value
        id_value = str(id_value).strip() if id_value is not None else f"row{i}"
        img_count = 1
        for j in range(col_from, col_to + 1):
            cell_value = ws.cell(row=i, column=j).value
            link = str(cell_value).strip() if cell_value is not None else ""
            if link and "http" in link:
                ext = guess_extension(link)
                file_name = f"{id_value}_image{img_count}{ext}"
                save_path = os.path.join(folder_path, file_name)
                tasks.append((link, save_path))
                img_count += 1
    return tasks


def download_file(url: str, save_path: str, timeout: int):
    try:
        resp = requests.get(url, timeout=timeout)
        if resp.status_code == 200:
            with open(save_path, "wb") as f:
                f.write(resp.content)
            return True, f"OK: {os.path.basename(save_path)}"
        return False, f"FAILED ({resp.status_code}): {url}"
    except requests.RequestException as e:
        return False, f"ERROR: {url} -> {e}"


def run_downloads(tasks, max_workers: int, timeout: int, progress_cb, log_cb):
    """
    progress_cb(done, total) is called after every completed download.
    log_cb(message) is called for every failed/error download.
    Returns (ok_count, failed_count).
    """
    ok = failed = done = 0
    total = len(tasks)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_task = {
            executor.submit(download_file, url, path, timeout): (url, path)
            for url, path in tasks
        }
        for future in as_completed(future_to_task):
            success, message = future.result()
            done += 1
            if success:
                ok += 1
            else:
                failed += 1
                log_cb(message)
            progress_cb(done, total)

    return ok, failed
